#!/usr/bin/python3

import openpyxl

wb_str = '20150927_Band Budget Analysis.xlsx'
sheet_str = 'pl_summary_md'
wb = openpyxl.load_workbook(wb_str)
sheet = wb.get_sheet_by_name(sheet_str)
start_cell = (2, 4)
end_cell = (7, 10)
for m in range(start_cell[1], end_cell[1] + 1):
    md_row = '|'
    for n in range(start_cell[0], end_cell[0] + 1):
        c = sheet.cell(row=m, column=n).value
        if c is None:
            c_str = ' '
            pass
        elif type(c) != str:
            c_str = '{:0.2f}'.format(c)
        else:
            c_str = c
#         md_row += str(sheet.cell(row=m, column=n).value)
        md_row += c_str
        md_row += '|'
#     md_row += '\n'
    if m == start_cell[1]:
        md_row += '\n' + '|' + "--|" * (end_cell[0] - start_cell[0] + 1)
    print(md_row)
